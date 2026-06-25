import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.utils import get_column_letter

random.seed(42)

PRODUCTOS = ["Waffle", "Banana Split", "Helado Soft", "Sundae", "Helado con Queso", "Tulipán con Queso"]
PRECIO_BASE = {"Waffle": 25, "Banana Split": 18, "Helado Soft": 20, "Sundae": 15, "Helado con Queso": 22, "Tulipán con Queso": 24}
ESTADOS = ["Nuevo Lead", "Contactado", "Interesado", "Propuesta", "Ganado", "Perdido"]
FUENTES = ["Facebook", "Instagram", "TikTok", "WhatsApp", "Referido", "Google"]
PRIORIDADES = ["Alta", "Media", "Baja"]
NOMBRES = ["María López", "Carlos Vera", "Andrea Torres", "Luis Vera", "María Zambrano", "Juan Tomalá",
           "Carla Reyes", "David Suárez", "Patricia González", "Jorge Mendoza", "Paola Vera", "Roberto Castillo",
           "Sofía López", "Kevin Rodríguez", "Daniela Cedeño", "Fernando Vega", "Isabel Navarro", "Tomás Herrera",
           "Camila Vargas", "Diego Salazar", "Natalia Cordero", "Eduardo Paredes", "Lucía Fonseca", "Mateo Bravo",
           "Joaquín Reyes", "Renata Morales", "Iván Cabrera", "Gabriela Ortiz", "Ricardo Salas", "Valentina Gómez"]
NOTAS = {
    "Nuevo Lead": ["Consultó precios", "Primera consulta", "Preguntó por el menú", "Escribió por redes"],
    "Contactado": ["Solicitó precios", "Se le envió catálogo", "Llamada de seguimiento realizada"],
    "Interesado": ["Interesado en promoción", "Pidió muestra", "Preguntó por pedidos grandes"],
    "Propuesta": ["Cotización enviada", "Esperando confirmación de pedido"],
    "Ganado": ["Compra realizada", "Pedido entregado", "Cliente satisfecho"],
    "Perdido": ["No respondió", "Eligió otra opción", "Precio fuera de presupuesto"],
}

MORADO = "6C2BD9"
COLOR_ESTADO = {
    "Nuevo Lead": "9AA3B2", "Contactado": "4FD1F0", "Interesado": "845EF7",
    "Propuesta": "F5A623", "Ganado": "2BB673", "Perdido": "E5484D",
}

wb = Workbook()
leads_ws = wb.active
leads_ws.title = "Leads"

# Columna A=ID, B=Fecha, C=Cliente, D=Telefono, E=Producto, F=Valor USD, G=Estado, H=Fuente, I=Prioridad, J=Notas, K=Fecha Seguimiento
headers = ["ID", "Fecha", "Cliente", "Telefono", "Producto", "Valor USD", "Estado", "Fuente", "Prioridad", "Notas", "Fecha Seguimiento"]
leads_ws.append(headers)
header_fill = PatternFill("solid", fgColor=MORADO)
for col_idx, _ in enumerate(headers, start=1):
    c = leads_ws.cell(row=1, column=col_idx)
    c.font = Font(name="Arial", bold=True, color="FFFFFF")
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

today = datetime.today()
N = 130
rows = []
for i in range(N):
    producto = random.choice(PRODUCTOS)
    estado = random.choice(ESTADOS)
    fecha = today - timedelta(days=random.randint(0, 120))
    valor = max(PRECIO_BASE[producto] + random.randint(-4, 6), 10)
    if estado not in ("Ganado", "Perdido"):
        seguimiento = fecha + timedelta(days=random.randint(2, 10))
        seguimiento_val = seguimiento.date()
    else:
        seguimiento_val = None
    rows.append([
        i + 1,
        fecha.date(),
        random.choice(NOMBRES),
        f"09{random.randint(10000000, 99999999)}",
        producto,
        valor,
        estado,
        random.choice(FUENTES),
        random.choice(PRIORIDADES),
        random.choice(NOTAS[estado]),
        seguimiento_val,
    ])

for r in rows:
    leads_ws.append(r)

last_row = N + 1
date_fmt = "yyyy-mm-dd"
for row in leads_ws.iter_rows(min_row=2, max_row=last_row, min_col=2, max_col=2):
    row[0].number_format = date_fmt
for row in leads_ws.iter_rows(min_row=2, max_row=last_row, min_col=11, max_col=11):
    row[0].number_format = date_fmt
for row in leads_ws.iter_rows(min_row=2, max_row=last_row, min_col=6, max_col=6):
    row[0].number_format = "$#,##0"
for row in leads_ws.iter_rows(min_row=2, max_row=last_row, min_col=4, max_col=4):
    row[0].number_format = "@"

widths = [6, 12, 22, 14, 18, 12, 14, 12, 10, 32, 16]
for i, w in enumerate(widths, start=1):
    leads_ws.column_dimensions[get_column_letter(i)].width = w

for col_letter, values in [("E", PRODUCTOS), ("G", ESTADOS), ("H", FUENTES), ("I", PRIORIDADES)]:
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    leads_ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")

for estado, hex_color in COLOR_ESTADO.items():
    font_color = "FFFFFF" if estado in ("Perdido", "Ganado", "Interesado") else "000000"
    rule = CellIsRule(operator="equal", formula=[f'"{estado}"'],
                       fill=PatternFill("solid", fgColor=hex_color),
                       font=Font(color=font_color, bold=True))
    leads_ws.conditional_formatting.add(f"G2:G{last_row}", rule)

leads_ws.freeze_panes = "A2"

# ---------------- Dashboard sheet ----------------
dash = wb.create_sheet("Dashboard")
dash["A1"] = "Fantasías Heladas CRM"
dash["A1"].font = Font(name="Arial", bold=True, size=18, color=MORADO)
dash["A2"] = "Resumen de clientes potenciales — valores en USD"
dash["A2"].font = Font(name="Arial", italic=True, size=11, color="6E6480")

kpi_labels = ["Total Leads", "Valor Pipeline", "Ingresos Ganados", "Tasa de Conversión"]
kpi_cells = ["B4", "D4", "F4", "H4"]
label_cells = ["B3", "D3", "F3", "H3"]
for label_cell, label in zip(label_cells, kpi_labels):
    dash[label_cell] = label
    dash[label_cell].font = Font(name="Arial", bold=True, size=10, color="6E6480")

dash["B4"] = "=COUNTA(Leads!C2:C131)"
dash["D4"] = '=SUMIFS(Leads!F2:F131,Leads!G2:G131,"<>Ganado",Leads!G2:G131,"<>Perdido")'
dash["F4"] = '=SUMIFS(Leads!F2:F131,Leads!G2:G131,"Ganado")'
dash["H4"] = '=COUNTIF(Leads!G2:G131,"Ganado")/(COUNTIF(Leads!G2:G131,"Ganado")+COUNTIF(Leads!G2:G131,"Perdido"))'

dash["D4"].number_format = "$#,##0"
dash["F4"].number_format = "$#,##0"
dash["H4"].number_format = "0.0%"
for cell in kpi_cells:
    dash[cell].font = Font(name="Arial", bold=True, size=20, color="2B1F3D")

# Helper tables -- placed far to the right (columns V+) so they never sit under the charts
dash["V1"] = "Por Estado"
dash["V1"].font = Font(bold=True, color=MORADO)
dash["V2"], dash["W2"] = "Estado", "Cantidad"
for i, estado in enumerate(ESTADOS):
    r = 3 + i
    dash[f"V{r}"] = estado
    dash[f"W{r}"] = f'=COUNTIF(Leads!G$2:G$131,V{r})'

dash["Y1"] = "Por Fuente"
dash["Y1"].font = Font(bold=True, color=MORADO)
dash["Y2"], dash["Z2"] = "Fuente", "Cantidad"
for i, fuente in enumerate(FUENTES):
    r = 3 + i
    dash[f"Y{r}"] = fuente
    dash[f"Z{r}"] = f'=COUNTIF(Leads!H$2:H$131,Y{r})'

dash["AB1"] = "Por Producto"
dash["AB1"].font = Font(bold=True, color=MORADO)
dash["AB2"], dash["AC2"] = "Producto", "Cantidad"
for i, producto in enumerate(PRODUCTOS):
    r = 3 + i
    dash[f"AB{r}"] = producto
    dash[f"AC{r}"] = f'=COUNTIF(Leads!E$2:E$131,AB{r})'

dash["AE1"] = "Por Mes"
dash["AE1"].font = Font(bold=True, color=MORADO)
dash["AE2"], dash["AF2"] = "Mes", "Cantidad"
months = []
cursor = today.replace(day=1)
for _ in range(6):
    months.append(cursor)
    prev_month = cursor.month - 1 or 12
    prev_year = cursor.year - 1 if cursor.month == 1 else cursor.year
    cursor = cursor.replace(year=prev_year, month=prev_month)
months.reverse()
for i, m in enumerate(months):
    r = 3 + i
    dash[f"AE{r}"] = m.strftime("%Y-%m")
    dash[f"AF{r}"] = f'=SUMPRODUCT((TEXT(Leads!B$2:B$131,"yyyy-mm")=AE{r})*1)'

for col in ("V", "W", "Y", "Z", "AB", "AC", "AE", "AF"):
    dash.column_dimensions[col].width = 16

# ---------------- Charts (2x2 grid, generous spacing so nothing overlaps) ----------------
def style_chart(chart, title):
    chart.title = title
    chart.height = 9
    chart.width = 15

estado_chart = DoughnutChart()
data = Reference(dash, min_col=23, min_row=2, max_row=8)   # W
cats = Reference(dash, min_col=22, min_row=3, max_row=8)   # V
estado_chart.add_data(data, titles_from_data=True)
estado_chart.set_categories(cats)
style_chart(estado_chart, "Embudo de ventas por estado")
dash.add_chart(estado_chart, "B7")

fuente_chart = DoughnutChart()
data = Reference(dash, min_col=26, min_row=2, max_row=8)   # Z
cats = Reference(dash, min_col=25, min_row=3, max_row=8)   # Y
fuente_chart.add_data(data, titles_from_data=True)
fuente_chart.set_categories(cats)
style_chart(fuente_chart, "Clientes potenciales por fuente")
dash.add_chart(fuente_chart, "N7")

producto_chart = BarChart()
producto_chart.type = "bar"
data = Reference(dash, min_col=29, min_row=2, max_row=8)   # AC
cats = Reference(dash, min_col=28, min_row=3, max_row=8)   # AB
producto_chart.add_data(data, titles_from_data=True)
producto_chart.set_categories(cats)
style_chart(producto_chart, "Clientes potenciales por producto")
dash.add_chart(producto_chart, "B26")

mes_chart = BarChart()
mes_chart.type = "col"
data = Reference(dash, min_col=32, min_row=2, max_row=8)   # AF
cats = Reference(dash, min_col=31, min_row=3, max_row=8)   # AE
mes_chart.add_data(data, titles_from_data=True)
mes_chart.set_categories(cats)
style_chart(mes_chart, "Nuevos clientes potenciales por mes")
dash.add_chart(mes_chart, "N26")

wb.save(r"C:\Users\Usuario\CRM\FantasiasHeladas_CRM_GoogleSheets.xlsx")
print("OK")
